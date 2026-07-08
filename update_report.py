# -*- coding: utf-8 -*-
"""
ระบบอัพเดทรายงานผลการฉีดวัคซีนไข้หวัดใหญ่ จังหวัดกาฬสินธุ์
================================================================
วิธีใช้: 
  1. วางไฟล์ rpt_data.csv ใหม่ลงใน folder เดียวกับ script นี้
  2. รัน: uv run --with openpyxl --no-project update_report.py
  3. ได้ไฟล์: 
     - ผลงานฉีดวัคซีนไข้หวัดใหญ่_กาฬสินธุ์.xlsx
     - Fludashboard.html
"""

import csv
import sys
import io
import os
import json
import subprocess
import shutil
from collections import defaultdict
from datetime import datetime

if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, 'rpt_data.csv')
EXCEL_OUTPUT = os.path.join(SCRIPT_DIR, 'ผลงานฉีดวัคซีนไข้หวัดใหญ่_กาฬสินธุ์.xlsx')
HTML_OUTPUT = os.path.join(SCRIPT_DIR, 'Fludashboard.html')

KALASIN_NETWORKS = {
    '10709': 'เมืองกาฬสินธุ์', '11077': 'นามน', '11078': 'กมลาไสย',
    '11079': 'ร่องคำ', '11080': 'เขาวง', '11081': 'ยางตลาด',
    '11082': 'ห้วยเม็ก', '11083': 'สหัสขันธ์', '11084': 'คำม่วง',
    '11085': 'ท่าคันโท', '11086': 'หนองกุงศรี', '11087': 'สมเด็จ',
    '11088': 'ห้วยผึ้ง', '11449': 'กุฉินารายณ์', '28017': 'นาคู',
    '28789': 'ฆ้องชัย', '28790': 'ดอนจาน', '28791': 'สามชัย',
}

DISTRICT_ORDER = [
    'เมืองกาฬสินธุ์', 'นามน', 'กมลาไสย', 'ร่องคำ', 'เขาวง',
    'ยางตลาด', 'ห้วยเม็ก', 'สหัสขันธ์', 'คำม่วง', 'ท่าคันโท',
    'หนองกุงศรี', 'สมเด็จ', 'ห้วยผึ้ง', 'กุฉินารายณ์',
    'นาคู', 'ฆ้องชัย', 'ดอนจาน', 'สามชัย',
]

# ============================================================
# DATA LOADING & PROCESSING
# ============================================================

def load_data(filepath):
    print(f"📂 กำลังอ่านไฟล์: {filepath}")
    with open(filepath, encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        all_rows = list(reader)
    kalasin_rows = [r for r in all_rows if len(r) >= 8 and r[1].strip() in KALASIN_NETWORKS]
    print(f"✅ พบข้อมูลกาฬสินธุ์: {len(kalasin_rows)} แถว")
    return kalasin_rows

def pivot_data(rows):
    units = {}
    for row in rows:
        net_code = row[1].strip()
        unit_type = row[2].strip()
        unit_code = row[3].strip()
        unit_name = row[4].strip()
        vaccine_type = row[5].strip()
        measure = row[6].strip()
        value = 0
        if row[7].strip():
            try: value = int(float(row[7].replace(',', '').strip()))
            except: value = 0
        if unit_code not in units:
            units[unit_code] = {
                'network': net_code, 'district': KALASIN_NETWORKS[net_code],
                'type': unit_type, 'name': unit_name,
                'v11_target_unit': 0, 'v11_target_total': 0, 'v11_vaccinated': 0,
                'v12_target_unit': 0, 'v12_target_total': 0, 'v12_vaccinated': 0,
            }
        is_v11 = 'V11' in vaccine_type
        is_v12 = 'V12' in vaccine_type
        if measure == 'ฉีด':
            if is_v11: units[unit_code]['v11_vaccinated'] = value
            elif is_v12: units[unit_code]['v12_vaccinated'] = value
        elif measure == 'เป้าหมาย(รายหน่วย)':
            if is_v11: units[unit_code]['v11_target_unit'] = value
            elif is_v12: units[unit_code]['v12_target_unit'] = value
        elif measure == 'เป้าหมาย(ทั้งหมด)':
            if is_v11: units[unit_code]['v11_target_total'] = value
            elif is_v12: units[unit_code]['v12_target_total'] = value
    return units

def aggregate_by_network(units):
    networks = defaultdict(lambda: {
        'district': '', 'main_name': '', 'main_code': '', 'sub_count': 0,
        'v11_target': 0, 'v11_vaccinated': 0, 'v12_target': 0, 'v12_vaccinated': 0, 'units': []
    })
    for code, info in units.items():
        net = info['network']
        networks[net]['district'] = info['district']
        if info['type'] == 'แม่ข่าย':
            networks[net]['main_name'] = info['name']
            networks[net]['main_code'] = code
            if info['v11_target_total'] > 0: networks[net]['v11_target'] = info['v11_target_total']
            if info['v12_target_total'] > 0: networks[net]['v12_target'] = info['v12_target_total']
        else:
            networks[net]['sub_count'] += 1
        networks[net]['v11_vaccinated'] += info['v11_vaccinated']
        networks[net]['v12_vaccinated'] += info['v12_vaccinated']
        networks[net]['units'].append(info | {'code': code})
    for net, data in networks.items():
        if data['v11_target'] == 0:
            data['v11_target'] = sum(u['v11_target_unit'] for u in data['units'])
        if data['v12_target'] == 0:
            data['v12_target'] = sum(u['v12_target_unit'] for u in data['units'])
    return dict(networks)

def aggregate_by_district(networks):
    districts = defaultdict(lambda: {
        'v11_target': 0, 'v11_vaccinated': 0, 'v12_target': 0, 'v12_vaccinated': 0,
        'unit_count': 0, 'network_codes': []
    })
    for net_code, data in networks.items():
        d = data['district']
        districts[d]['v11_target'] += data['v11_target']
        districts[d]['v11_vaccinated'] += data['v11_vaccinated']
        districts[d]['v12_target'] += data['v12_target']
        districts[d]['v12_vaccinated'] += data['v12_vaccinated']
        districts[d]['unit_count'] += 1 + data['sub_count']
        districts[d]['network_codes'].append(net_code)
    return dict(districts)

def safe_pct(a, b):
    return round((a / b) * 100, 2) if b > 0 else 0.0

# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(units, networks, districts, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    hf = Font(name='TH Sarabun New', size=14, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    df = Font(name='TH Sarabun New', size=14)
    dfb = Font(name='TH Sarabun New', size=14, bold=True)
    tf = Font(name='TH Sarabun New', size=18, bold=True, color='1F4E79')
    sf = Font(name='TH Sarabun New', size=14, italic=True, color='666666')
    stf = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    gtf = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yf = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def style_header(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = tb

    def style_row(ws, row, ncols, font=None, fill=None, aligns=None):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = font or df; cell.border = tb
            if fill: cell.fill = fill
            cell.alignment = (aligns[c-1] if aligns and c <= len(aligns) else ca)

    def pct_color(cell, pct):
        if pct >= 80: cell.fill = gf
        elif pct >= 50: cell.fill = yf
        else: cell.fill = rf

    utime = datetime.now().strftime('%d/%m/%Y %H:%M น.')

    # === SHEET 1 ===
    ws1 = wb.active; ws1.title = 'สรุปรายอำเภอ'; ws1.sheet_properties.tabColor = '1F4E79'
    ws1.merge_cells('A1:L1')
    ws1['A1'] = 'ผลงานการฉีดวัคซีนป้องกันโรคไข้หวัดใหญ่ตามฤดูกาล จังหวัดกาฬสินธุ์'
    ws1['A1'].font = tf; ws1['A1'].alignment = ca
    ws1.merge_cells('A2:L2'); ws1['A2'] = f'ข้อมูล ณ วันที่ {utime}'; ws1['A2'].font = sf; ws1['A2'].alignment = ca
    h1 = ['ลำดับ','อำเภอ','จำนวน\nหน่วยบริการ','เป้าหมาย\nV11','ฉีด\nV11','% V11','เป้าหมาย\nV12','ฉีด\nV12','% V12','เป้าหมายรวม','ฉีดรวม','% รวม']
    r = 4
    for c, h in enumerate(h1, 1): ws1.cell(row=r, column=c, value=h)
    style_header(ws1, r, len(h1))
    r = 5; gv11t = gv11v = gv12t = gv12v = 0
    for idx, dn in enumerate(DISTRICT_ORDER, 1):
        d = districts.get(dn, {'v11_target':0,'v11_vaccinated':0,'v12_target':0,'v12_vaccinated':0,'unit_count':0})
        p1 = safe_pct(d['v11_vaccinated'], d['v11_target'])
        p2 = safe_pct(d['v12_vaccinated'], d['v12_target'])
        tt = d['v11_target']+d['v12_target']; tv = d['v11_vaccinated']+d['v12_vaccinated']
        pt = safe_pct(tv, tt)
        vals = [idx, dn, d['unit_count'], d['v11_target'], d['v11_vaccinated'], p1, d['v12_target'], d['v12_vaccinated'], p2, tt, tv, pt]
        for c, v in enumerate(vals, 1): ws1.cell(row=r, column=c, value=v)
        al = [ca, la] + [ca]*10
        style_row(ws1, r, len(vals), aligns=al)
        pct_color(ws1.cell(row=r, column=6), p1); pct_color(ws1.cell(row=r, column=9), p2); pct_color(ws1.cell(row=r, column=12), pt)
        for pc in [6,9,12]: ws1.cell(row=r, column=pc).number_format = '0.00"%"'
        gv11t += d['v11_target']; gv11v += d['v11_vaccinated']; gv12t += d['v12_target']; gv12v += d['v12_vaccinated']
        r += 1
    gtt = gv11t+gv12t; gtv = gv11v+gv12v
    tu = sum(d.get('unit_count',0) for d in districts.values())
    gvals = ['','รวมทั้งจังหวัด',tu,gv11t,gv11v,safe_pct(gv11v,gv11t),gv12t,gv12v,safe_pct(gv12v,gv12t),gtt,gtv,safe_pct(gtv,gtt)]
    for c, v in enumerate(gvals, 1): ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, len(gvals), font=Font(name='TH Sarabun New', size=14, bold=True, color='FFFFFF'), fill=gtf)
    for pc in [6,9,12]: ws1.cell(row=r, column=pc).number_format = '0.00"%"'
    for i, w in enumerate([8, 25, 14, 14, 12, 12, 14, 12, 12, 16, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A5'; ws1.auto_filter.ref = f'A4:{get_column_letter(len(h1))}{r}'
    chart = BarChart(); chart.type = 'col'; chart.title = 'เปรียบเทียบเป้าหมายและผลงาน V11 รายอำเภอ'
    chart.y_axis.title = 'จำนวน (ราย)'; chart.style = 10; chart.width = 35; chart.height = 15
    dref = Reference(ws1, min_col=4, min_row=4, max_col=5, max_row=4+len(DISTRICT_ORDER))
    cats = Reference(ws1, min_col=2, min_row=5, max_row=4+len(DISTRICT_ORDER))
    chart.add_data(dref, titles_from_data=True); chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].graphicalProperties.solidFill = "70AD47"
    ws1.add_chart(chart, f'A{r+3}')

    # === SHEET 2 ===
    ws2 = wb.create_sheet('รายเครือข่าย'); ws2.sheet_properties.tabColor = '2E75B6'
    ws2.merge_cells('A1:N1')
    ws2['A1'] = 'ผลงานรายเครือข่าย (แม่ข่าย+ลูกข่าย) จังหวัดกาฬสินธุ์'
    ws2['A1'].font = tf; ws2['A1'].alignment = ca
    ws2.merge_cells('A2:N2'); ws2['A2'] = f'ข้อมูล ณ วันที่ {utime}'; ws2['A2'].font = sf; ws2['A2'].alignment = ca
    h2 = ['ลำดับ','อำเภอ','เครือข่าย (แม่ข่าย)','รหัส','จำนวน\nลูกข่าย','เป้าหมาย\nV11','ฉีด\nV11','% V11','เป้าหมาย\nV12','ฉีด\nV12','% V12','เป้าหมายรวม','ฉีดรวม','% รวม']
    r = 4
    for c, h in enumerate(h2, 1): ws2.cell(row=r, column=c, value=h)
    style_header(ws2, r, len(h2))
    r = 5; idx = 0
    snets = sorted(networks.items(), key=lambda x: (DISTRICT_ORDER.index(x[1]['district']) if x[1]['district'] in DISTRICT_ORDER else 99, x[0]))
    for nc, data in snets:
        idx += 1
        p1 = safe_pct(data['v11_vaccinated'], data['v11_target']); p2 = safe_pct(data['v12_vaccinated'], data['v12_target'])
        tt = data['v11_target']+data['v12_target']; tv = data['v11_vaccinated']+data['v12_vaccinated']; pt = safe_pct(tv, tt)
        vals = [idx, data['district'], data['main_name'], nc, data['sub_count'], data['v11_target'], data['v11_vaccinated'], p1, data['v12_target'], data['v12_vaccinated'], p2, tt, tv, pt]
        for c, v in enumerate(vals, 1): ws2.cell(row=r, column=c, value=v)
        al = [ca, la, la, ca] + [ca]*10
        style_row(ws2, r, len(vals), aligns=al)
        pct_color(ws2.cell(row=r, column=8), p1); pct_color(ws2.cell(row=r, column=11), p2); pct_color(ws2.cell(row=r, column=14), pt)
        for pc in [8,11,14]: ws2.cell(row=r, column=pc).number_format = '0.00"%"'
        r += 1
    gv2 = ['','รวมทั้งจังหวัด','','',tu-18,gv11t,gv11v,safe_pct(gv11v,gv11t),gv12t,gv12v,safe_pct(gv12v,gv12t),gtt,gtv,safe_pct(gtv,gtt)]
    for c, v in enumerate(gv2, 1): ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, len(gv2), font=Font(name='TH Sarabun New', size=14, bold=True, color='FFFFFF'), fill=gtf)
    for pc in [8,11,14]: ws2.cell(row=r, column=pc).number_format = '0.00"%"'
    for i, w in enumerate([8, 22, 36, 10, 12, 14, 12, 12, 14, 12, 12, 16, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A5'; ws2.auto_filter.ref = f'A4:{get_column_letter(len(h2))}{r}'

    # === SHEET 3 ===
    ws3 = wb.create_sheet('รายหน่วยบริการ'); ws3.sheet_properties.tabColor = '548235'
    ws3.merge_cells('A1:L1')
    ws3['A1'] = 'รายละเอียดผลงานทุกหน่วยบริการ จังหวัดกาฬสินธุ์'
    ws3['A1'].font = tf; ws3['A1'].alignment = ca
    ws3.merge_cells('A2:L2'); ws3['A2'] = f'ข้อมูล ณ วันที่ {utime}'; ws3['A2'].font = sf; ws3['A2'].alignment = ca
    h3 = ['ลำดับ','อำเภอ','เครือข่าย','ประเภท','รหัส','ชื่อหน่วยบริการ','เป้าหมาย\nV11','ฉีด\nV11','% V11','เป้าหมาย\nV12','ฉีด\nV12','% V12']
    r = 4
    for c, h in enumerate(h3, 1): ws3.cell(row=r, column=c, value=h)
    style_header(ws3, r, len(h3))
    r = 5; idx = 0
    su = sorted(units.items(), key=lambda x: (DISTRICT_ORDER.index(x[1]['district']) if x[1]['district'] in DISTRICT_ORDER else 99, x[1]['network'], 0 if x[1]['type']=='แม่ข่าย' else 1, x[0]))
    for code, info in su:
        idx += 1
        vt1 = info['v11_target_unit']; vt2 = info['v12_target_unit']
        p1 = safe_pct(info['v11_vaccinated'], vt1); p2 = safe_pct(info['v12_vaccinated'], vt2)
        vals = [idx, info['district'], info['network'], info['type'], code, info['name'], vt1, info['v11_vaccinated'], p1, vt2, info['v12_vaccinated'], p2]
        for c, v in enumerate(vals, 1): ws3.cell(row=r, column=c, value=v)
        ism = info['type'] == 'แม่ข่าย'
        al = [ca, la, ca, ca, ca, la] + [ca]*6
        style_row(ws3, r, len(vals), font=dfb if ism else df, fill=stf if ism else None, aligns=al)
        if vt1 > 0: pct_color(ws3.cell(row=r, column=9), p1)
        if vt2 > 0: pct_color(ws3.cell(row=r, column=12), p2)
        for pc in [9,12]: ws3.cell(row=r, column=pc).number_format = '0.00"%"'
        r += 1
    for i, w in enumerate([8, 22, 12, 12, 10, 50, 14, 12, 12, 14, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A5'; ws3.auto_filter.ref = f'A4:{get_column_letter(len(h3))}{r-1}'
    wb.save(output_path)
    print(f"📊 สร้างไฟล์ Excel เรียบร้อย: {output_path}")

# ============================================================
# DASHBOARD HTML
# ============================================================

def generate_dashboard(units, networks, districts, output_path):
    utime = datetime.now().strftime('%d/%m/%Y %H:%M น.')
    gv11t = sum(d['v11_target'] for d in districts.values())
    gv11v = sum(d['v11_vaccinated'] for d in districts.values())
    gv12t = sum(d['v12_target'] for d in districts.values())
    gv12v = sum(d['v12_vaccinated'] for d in districts.values())
    gtt = gv11t+gv12t; gtv = gv11v+gv12v; gp = safe_pct(gtv, gtt)
    tu = sum(d['unit_count'] for d in districts.values())

    dl = []; dv11t = []; dv11v = []; dv12t = []; dv12v = []; dpct = []
    for dn in DISTRICT_ORDER:
        d = districts.get(dn, {'v11_target':0,'v11_vaccinated':0,'v12_target':0,'v12_vaccinated':0,'unit_count':0})
        dl.append(dn); dv11t.append(d['v11_target']); dv11v.append(d['v11_vaccinated'])
        dv12t.append(d['v12_target']); dv12v.append(d['v12_vaccinated'])
        dpct.append(safe_pct(d['v11_vaccinated']+d['v12_vaccinated'], d['v11_target']+d['v12_target']))

    # Sorted pct data for horizontal bar
    sorted_idx = sorted(range(len(dpct)), key=lambda i: dpct[i], reverse=True)
    sorted_labels = [dl[i] for i in sorted_idx]
    sorted_pcts = [dpct[i] for i in sorted_idx]

    ndata = []
    snets = sorted(networks.items(), key=lambda x: (DISTRICT_ORDER.index(x[1]['district']) if x[1]['district'] in DISTRICT_ORDER else 99, x[0]))
    for nc, data in snets:
        tt = data['v11_target']+data['v12_target']; tv = data['v11_vaccinated']+data['v12_vaccinated']
        subs = []
        for u in sorted(data['units'], key=lambda u: (0 if u['type']=='แม่ข่าย' else 1, u['code'])):
            ut = u['v11_target_unit']+u['v12_target_unit']; uv = u['v11_vaccinated']+u['v12_vaccinated']
            subs.append({'code':u['code'],'name':u['name'],'type':u['type'],'v11t':u['v11_target_unit'],'v11v':u['v11_vaccinated'],'v12t':u['v12_target_unit'],'v12v':u['v12_vaccinated'],'tt':ut,'tv':uv,'pct':safe_pct(uv,ut)})
        ndata.append({'nc':nc,'district':data['district'],'mn':data['main_name'],'sc':data['sub_count'],'v11t':data['v11_target'],'v11v':data['v11_vaccinated'],'v12t':data['v12_target'],'v12v':data['v12_vaccinated'],'tt':tt,'tv':tv,'pct':safe_pct(tv,tt),'units':subs})

    nj = json.dumps(ndata, ensure_ascii=False)

    district_pcts = []
    for dn in DISTRICT_ORDER:
        d = districts.get(dn, {'v11_target':0,'v11_vaccinated':0,'v12_target':0,'v12_vaccinated':0})
        tt = d['v11_target']+d['v12_target']
        tv = d['v11_vaccinated']+d['v12_vaccinated']
        pct = safe_pct(tv, tt)
        district_pcts.append((dn, pct))
    
    # Sort from high to low percentage
    sorted_district_pcts = sorted(district_pcts, key=lambda x: x[1], reverse=True)
    
    district_cards = []
    for dn, pct in sorted_district_pcts:
        pctClass = 'ph' if pct >= 80 else 'pm' if pct >= 50 else 'pl'
        card_html = f'<div class="da-grid-card" id="dg-{dn}" onclick="selectDistrictCard(\'{dn}\')"><div class="dg-name">{dn}</div><div class="dg-pct {pctClass}">{pct:.1f}%</div></div>'
        district_cards.append(card_html)
    grid_html = '\n'.join(district_cards)

    html = f'''<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard ฉีดวัคซีนไข้หวัดใหญ่ กาฬสินธุ์</title>
<meta name="description" content="แดชบอร์ดสรุปผลการฉีดวัคซีนป้องกันโรคไข้หวัดใหญ่ตามฤดูกาล จังหวัดกาฬสินธุ์ เขตสุขภาพที่ 7">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+Thai:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<style>
:root {{
  --bg:#0f172a;
  --bg2:#1e293b;
  --bg3:#334155;
  --fg:#f1f5f9;
  --fg2:#94a3b8;
  --fg3:#64748b;
  --blue:#3b82f6;
  --cyan:#06b6d4;
  --green:#10b981;
  --amber:#f59e0b;
  --red:#ef4444;
  --purple:#8b5cf6;
  --bdr:#334155;
  --r:16px;
}}
body.theme-light {{
  --bg:#f8fafc;
  --bg2:#ffffff;
  --bg3:#f1f5f9;
  --fg:#0f172a;
  --fg2:#475569;
  --fg3:#94a3b8;
  --bdr:#cbd5e1;
}}
body.theme-blue {{
  --bg:#0b132b;
  --bg2:#1c2541;
  --bg3:#3a506b;
  --fg:#ffffff;
  --fg2:#8b9bb4;
  --fg3:#5c6b8c;
  --blue:#00b4d8;
  --cyan:#90e0ef;
  --bdr:#3a506b;
}}
body.theme-emerald {{
  --bg:#022c22;
  --bg2:#064e3b;
  --bg3:#0f766e;
  --fg:#f0fdf4;
  --fg2:#a7f3d0;
  --fg3:#34d399;
  --blue:#10b981;
  --cyan:#34d399;
  --bdr:#0f766e;
}}

/* Theme Selector */
.theme-selector {{
  display: flex;
  gap: 8px;
  align-items: center;
  margin-left: auto;
  padding-left: 16px;
}}
.theme-btn {{
  width: 20px;
  height: 20px;
  border-radius: 50%;
  border: 2px solid var(--bdr);
  cursor: pointer;
  transition: transform 0.2s, border-color 0.2s;
  position: relative;
  outline: none;
}}
.theme-btn:hover {{
  transform: scale(1.2);
}}
.theme-btn.active {{
  border-color: var(--fg);
  transform: scale(1.1);
  box-shadow: 0 0 8px var(--blue);
}}
.tb-dark {{ background: #0f172a; }}
.tb-light {{ background: #ffffff; border-color: #cbd5e1; }}
.tb-blue {{ background: #00b4d8; }}
.tb-emerald {{ background: #10b981; }}

*{{margin:0;padding:0;box-sizing:border-box;}}
html{{scroll-behavior:smooth;}}
body{{font-family:'Noto Sans Thai',sans-serif;background:var(--bg);color:var(--fg);line-height:1.6;transition:background 0.3s, color 0.3s;}}
.container{{max-width:1400px;margin:0 auto;padding:24px;}}

/* Loader */
#loader{{position:fixed;top:0;left:0;width:100%;height:100%;background:var(--bg);display:flex;align-items:center;justify-content:center;z-index:9999;transition:opacity .5s;}}
#loader.hide{{opacity:0;pointer-events:none;}}
.spinner{{width:48px;height:48px;border:4px solid var(--bdr);border-top-color:var(--blue);border-radius:50%;animation:spin .8s linear infinite;}}
@keyframes spin{{to{{transform:rotate(360deg);}}}}

/* Nav */
.nav{{position:sticky;top:0;z-index:100;background:rgba(15,23,42,.9);backdrop-filter:blur(12px);border-bottom:1px solid var(--bdr);padding:0 24px;}}
.nav-inner{{max-width:1400px;margin:0 auto;display:flex;align-items:center;gap:8px;overflow-x:auto;padding:8px 0;}}
.nav a{{color:var(--fg2);text-decoration:none;padding:8px 16px;border-radius:8px;font-size:.85rem;font-weight:500;white-space:nowrap;transition:all .2s;}}
.nav a:hover,.nav a.active{{color:var(--blue);background:rgba(59,130,246,.1);}}

/* Header */
.header{{text-align:center;padding:32px 0 24px;position:relative;}}
.header::after{{content:'';position:absolute;bottom:0;left:50%;transform:translateX(-50%);width:120px;height:3px;background:linear-gradient(135deg,var(--blue),var(--cyan));border-radius:2px;}}
.header h1{{font-size:2.4rem;font-weight:800;background:linear-gradient(135deg,var(--blue),var(--cyan));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:8px;line-height:1.3;}}
.header .sub{{color:var(--fg);font-size:1.2rem;font-weight:600;margin-top:6px;}}
.badge{{display:inline-block;margin-top:14px;padding:6px 20px;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.3);border-radius:999px;font-size:0.9rem;font-weight:600;color:var(--blue);}}

/* KPI */
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin:32px 0;}}
.kpi{{background:var(--bg2);border:1px solid var(--bdr);border-radius:var(--r);padding:24px;position:relative;overflow:hidden;transition:transform .3s,box-shadow .3s;}}
.kpi:hover{{transform:translateY(-4px);box-shadow:0 10px 25px rgba(0,0,0,.4);}}
.kpi::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;}}
.kpi:nth-child(1)::before{{background:linear-gradient(135deg,var(--blue),var(--cyan));}}
.kpi:nth-child(2)::before{{background:linear-gradient(135deg,var(--green),var(--cyan));}}
.kpi:nth-child(3)::before{{background:linear-gradient(135deg,var(--amber),#f97316);}}
.kpi:nth-child(4)::before{{background:linear-gradient(135deg,var(--purple),#ec4899);}}
.kpi:nth-child(5)::before{{background:linear-gradient(135deg,var(--red),#f97316);}}
.kpi-label{{font-size:.85rem;color:var(--fg2);margin-bottom:8px;font-weight:500;}}
.kpi-val{{font-size:2rem;font-weight:700;line-height:1.2;}}
.kpi:nth-child(1) .kpi-val{{color:var(--blue);}}
.kpi:nth-child(2) .kpi-val{{color:var(--green);}}
.kpi:nth-child(3) .kpi-val{{color:var(--amber);}}
.kpi:nth-child(4) .kpi-val{{color:var(--purple);}}
.kpi:nth-child(5) .kpi-val{{color:var(--red);}}
.kpi-sub{{font-size:.8rem;color:var(--fg3);margin-top:4px;}}
.pbar{{width:100%;height:8px;background:var(--bg);border-radius:4px;overflow:hidden;margin-top:4px;}}
.pfill{{height:100%;border-radius:4px;transition:width 1.5s ease;}}

/* Cards */
.card{{background:var(--bg2);border:1px solid var(--bdr);border-radius:var(--r);padding:24px;margin:24px 0;}}
.card h3{{font-size:1.1rem;font-weight:600;margin-bottom:16px;}}
.chart-box{{position:relative;width:100%;height:400px;}}

/* Tabs */
.tabs{{display:flex;gap:4px;padding:4px;background:var(--bg);border-radius:12px;width:fit-content;}}
.tab{{padding:8px 20px;border:none;background:transparent;color:var(--fg2);font-family:'Noto Sans Thai',sans-serif;font-size:.9rem;font-weight:500;border-radius:8px;cursor:pointer;transition:all .2s;}}
.tab.active{{background:var(--blue);color:#fff;}}
.tab:hover:not(.active){{background:var(--bg3);color:var(--fg);}}

/* Table */
.tbl-header{{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;margin-bottom:16px;}}
.search{{padding:8px 16px;border:1px solid var(--bdr);border-radius:8px;background:var(--bg);color:var(--fg);font-family:'Noto Sans Thai',sans-serif;font-size:.9rem;width:260px;outline:none;transition:border-color .2s;}}
.search:focus{{border-color:var(--blue);}}
.btn{{padding:8px 16px;border:1px solid var(--bdr);border-radius:8px;background:var(--bg);color:var(--fg2);font-family:'Noto Sans Thai',sans-serif;font-size:.85rem;cursor:pointer;transition:all .2s;display:inline-flex;align-items:center;gap:6px;}}
.btn:hover{{border-color:var(--blue);color:var(--blue);background:rgba(59,130,246,.1);}}
.twrap{{overflow-x:auto;}}
table{{width:100%;border-collapse:collapse;font-size:.9rem;}}
thead th{{background:var(--bg);color:var(--fg2);font-weight:600;padding:12px 16px;text-align:center;white-space:nowrap;position:sticky;top:0;cursor:pointer;user-select:none;transition:color .2s;}}
thead th:hover{{color:var(--blue);}}
thead th.sa::after{{content:' ▲';font-size:.7em;}}
thead th.sd::after{{content:' ▼';font-size:.7em;}}
tbody td{{padding:10px 16px;border-bottom:1px solid var(--bdr);text-align:center;white-space:nowrap;}}
tbody td:nth-child(2){{text-align:left;}}
tbody tr{{transition:background .15s;}}
tbody tr:hover{{background:var(--bg3);}}
tbody tr.mu{{background:rgba(59,130,246,.08);font-weight:600;}}
.pct-b{{display:inline-block;padding:2px 10px;border-radius:999px;font-size:.8rem;font-weight:600;}}
.ph{{background:rgba(16,185,129,.15);color:#10b981;}}
.pm{{background:rgba(245,158,11,.15);color:#f59e0b;}}
.pl{{background:rgba(239,68,68,.15);color:#ef4444;}}
.eb{{background:none;border:1px solid var(--bdr);color:var(--fg2);width:28px;height:28px;border-radius:6px;cursor:pointer;font-size:.8rem;display:inline-flex;align-items:center;justify-content:center;transition:all .2s;}}
.eb:hover{{border-color:var(--blue);color:var(--blue);background:rgba(59,130,246,.1);}}
.eb.ex{{background:var(--blue);color:#fff;border-color:var(--blue);}}
.sr{{display:none;}}
.sr.show{{display:table-row;}}
.sr td{{background:rgba(0,0,0,.15);font-size:.85rem;padding:6px 16px;color:var(--fg2);}}
.sr td:nth-child(2){{padding-left:36px;}}

/* District Analysis */
.da-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0;}}
.da-kpi{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin:16px 0;}}
.da-kpi-card{{background:var(--bg);border-radius:12px;padding:16px;text-align:center;border:1px solid var(--bdr);box-shadow:inset 0 1px 3px rgba(0,0,0,0.2);}}
.da-kpi-card .val{{font-size:1.6rem;font-weight:700;}}
.da-kpi-card .lbl{{font-size:.75rem;color:var(--fg3);margin-top:4px;}}

/* Responsive District Grid */
.da-district-grid {{
  display: grid;
  grid-template-columns: repeat(6, 1fr);
  gap: 10px;
  margin: 16px 0 24px 0;
}}
.da-grid-card {{
  background: rgba(30, 41, 59, 0.4);
  border: 1px solid var(--bdr);
  border-radius: 8px;
  padding: 10px 14px;
  cursor: pointer;
  transition: all 0.2s ease-in-out;
  display: flex;
  flex-direction: row;
  justify-content: space-between;
  align-items: center;
  gap: 6px;
  box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
  height: 48px;
  box-sizing: border-box;
}}
.da-grid-card:hover {{
  transform: translateY(-1px);
  border-color: var(--blue);
  background: rgba(30, 41, 59, 0.8);
  box-shadow: 0 4px 8px rgba(59, 130, 246, 0.15);
}}
.da-grid-card.active {{
  background: linear-gradient(135deg, var(--blue), var(--cyan));
  border-color: transparent;
  box-shadow: 0 4px 12px rgba(6, 182, 212, 0.3);
}}
.dg-name {{
  font-size: 1.05rem;
  font-weight: 600;
  color: var(--fg);
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  text-align: left;
  margin-bottom: 0 !important;
  flex-grow: 1;
  transition: color 0.2s;
}}
.da-grid-card:hover .dg-name {{
  color: var(--fg);
}}
.da-grid-card.active .dg-name {{
  color: #fff !important;
  font-weight: 700;
}}
.dg-pct {{
  font-size: 0.85rem;
  font-weight: 700;
  padding: 3px 10px;
  border-radius: 6px;
  white-space: nowrap;
  transition: all 0.2s;
}}
.da-grid-card.active .dg-pct {{
  background: rgba(255, 255, 255, 0.2) !important;
  color: #fff !important;
}}

@media(max-width: 991px) {{
  .da-district-grid {{
    grid-template-columns: repeat(3, 1fr) !important;
  }}
}}
@media(max-width: 480px) {{
  .da-district-grid {{
    grid-template-columns: repeat(2, 1fr) !important;
  }}
}}

/* Dynamic District Banner */
.selected-district-banner {{
  background: linear-gradient(135deg, rgba(59, 130, 246, 0.12), rgba(6, 182, 212, 0.04));
  border-left: 5px solid var(--blue);
  padding: 18px 24px;
  border-radius: 12px;
  margin: 16px 0 24px 0;
  box-shadow: 0 4px 20px rgba(0,0,0,0.2);
  border: 1px solid rgba(59, 130, 246, 0.15);
  border-left: 5px solid var(--blue);
  animation: fadeIn 0.4s ease-out;
}}
@keyframes fadeIn {{
  from {{ opacity: 0; transform: translateY(10px); }}
  to {{ opacity: 1; transform: translateY(0); }}
}}
.banner-badge {{
  font-size: 0.75rem;
  text-transform: uppercase;
  letter-spacing: 1px;
  color: var(--cyan);
  font-weight: 700;
  display: block;
  margin-bottom: 4px;
}}
.banner-title {{
  font-size: 1.8rem;
  font-weight: 700;
  color: var(--fg);
  text-shadow: 0 0 10px rgba(59, 130, 246, 0.2);
}}

/* Scroll top */
#scrollTop{{position:fixed;bottom:24px;right:24px;width:44px;height:44px;border-radius:50%;background:var(--blue);color:#fff;border:none;font-size:1.2rem;cursor:pointer;display:none;align-items:center;justify-content:center;box-shadow:0 4px 15px rgba(59,130,246,.4);transition:all .3s;z-index:90;}}
#scrollTop:hover{{transform:scale(1.1);}}
#scrollTop.show{{display:flex;}}

/* Footer */
.footer{{text-align:center;padding:32px 0 16px;color:var(--fg3);font-size:.8rem;}}

@media(max-width:768px){{
.container{{padding:12px;}}
.header h1{{font-size:1.6rem;}}
.kpi-grid{{grid-template-columns:repeat(2,1fr);}}
.kpi-val{{font-size:1.5rem;}}
.chart-box{{height:300px;}}
.search{{width:100%;}}
.da-grid{{grid-template-columns:1fr;}}
.sel{{min-width:100%;}}
.da-grid-card {{
  height: 44px !important;
  padding: 8px 10px !important;
}}
.dg-name {{
  font-size: 0.9rem !important;
}}
.dg-pct {{
  font-size: 0.78rem !important;
  padding: 2px 6px !important;
}}
}}
@media(max-width:480px){{
.kpi-grid{{grid-template-columns:1fr;}}
.da-grid-card {{
  height: 40px !important;
  padding: 6px 8px !important;
}}
.dg-name {{
  font-size: 0.8rem !important;
}}
.dg-pct {{
  font-size: 0.7rem !important;
  padding: 1px 4px !important;
}}
}}
</style>
</head>
<body>

<div id="loader"><div class="spinner"></div></div>

<nav class="nav" id="nav-bar">
<div class="nav-inner">
<a href="#district-selector">🔍 เลือกพื้นที่</a>
<a href="#kpi-section">📊 สรุปผลงาน</a>
<a href="#charts-section">📈 กราฟวิเคราะห์</a>
<a href="#table-section">📋 ตารางข้อมูล</a>
<div class="theme-selector">
  <button class="theme-btn tb-dark active" data-theme="dark" onclick="setTheme('dark')" title="ธีมมืด (Dark)"></button>
  <button class="theme-btn tb-light" data-theme="light" onclick="setTheme('light')" title="ธีมสว่าง (Light)"></button>
  <button class="theme-btn tb-blue" data-theme="blue" onclick="setTheme('blue')" title="ธีมสีฟ้า (Ocean)"></button>
  <button class="theme-btn tb-emerald" data-theme="emerald" onclick="setTheme('emerald')" title="ธีมเขียวมรกต (Emerald)"></button>
</div>
</div>
</nav>

<div class="container">
<header class="header">
<h1>💉 ผลงานฉีดวัคซีนไข้หวัดใหญ่ตามฤดูกาลของประชาชนกลุ่มเสี่ยง ปี 2569</h1>
<p class="sub">จังหวัดกาฬสินธุ์ · เขตสุขภาพที่ 7</p>
<span class="badge">🕐 อัพเดท: {utime}</span>
</header>

<div class="card" id="district-selector" style="margin-top: 0;">
<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;margin-bottom:16px;">
<div>
<h3 style="margin-bottom:4px;">🔍 เลือกพื้นที่ที่ต้องการดูข้อมูล</h3>
<p style="color:var(--fg3);font-size:0.85rem;">คลิกเลือกปุ่มภาพรวมจังหวัด หรือรายอำเภอในตาราง Grid ด้านล่าง เพื่อดูข้อมูลของพื้นที่นั้น</p>
</div>
<button class="btn" onclick="exportCSV('district')">📥 ส่งออก CSV ข้อมูลที่เลือก</button>
</div>

<div style="display:flex;justify-content:center;margin-bottom:16px;">
  <div class="da-grid-card active" id="dg-ภาพรวมจังหวัด" onclick="selectDistrictCard('ภาพรวมจังหวัด')" style="width: 100%; max-width: 340px; justify-content: center; gap: 12px; height: 48px;">
    <div class="dg-name" style="flex-grow:0; font-size: 1.15rem; font-weight: 700;">📍 ภาพรวมทั้งจังหวัด</div>
    <div class="dg-pct ph" style="font-size: 0.95rem; padding: 4px 12px;">{gp:.1f}%</div>
  </div>
</div>

<div class="da-district-grid" id="daDistrictGrid">
{grid_html}
</div>
</div>

<div class="selected-district-banner" id="daDistrictBanner">
<span class="banner-badge">📍 พื้นที่รายงานข้อมูล</span>
<h2 class="banner-title" id="daDistrictName">ภาพรวมจังหวัดกาฬสินธุ์</h2>
</div>

<div class="kpi-grid" id="kpi-section">
<div class="kpi"><div class="kpi-label">🎯 เป้าหมายรวม</div><div class="kpi-val" id="kpi-tt" data-count="0">0</div><div class="kpi-sub" id="kpi-sub-t">V11: 0 · V12: 0</div></div>
<div class="kpi"><div class="kpi-label">💉 ฉีดแล้ว</div><div class="kpi-val" id="kpi-tv" data-count="0">0</div><div class="kpi-sub" id="kpi-sub-v">V11: 0 · V12: 0</div></div>
<div class="kpi"><div class="kpi-label">📊 ร้อยละผลงาน</div><div class="kpi-val" id="kpi-pct" data-count="0" data-suffix="%">0%</div><div class="pbar"><div class="pfill" id="kpi-pfill" style="width:0%;"></div></div></div>
<div class="kpi"><div class="kpi-label" id="kpi-lbl-units">🏥 อำเภอ</div><div class="kpi-val" id="kpi-units" data-count="0">0</div><div class="kpi-sub" id="kpi-sub-units">อำเภอ</div></div>
<div class="kpi"><div class="kpi-label" id="kpi-lbl-total-units">🏠 หน่วยบริการ</div><div class="kpi-val" id="kpi-total-units" data-count="0">0</div><div class="kpi-sub" id="kpi-sub-total-units">แม่ข่าย+ลูกข่าย</div></div>
</div>

<div id="charts-section">
  <div id="provincial-charts">
    <div class="card" id="bar-chart-card">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;">
    <h3>📊 เปรียบเทียบผลงานรายอำเภอ</h3>
    <div class="tabs"><button class="tab active" onclick="swChart('v11')" id="tv11">V11 (7กลุ่มเสี่ยง)</button><button class="tab" onclick="swChart('v12')" id="tv12">V12 (กลุ่มพิเศษ)</button><button class="tab" onclick="swChart('total')" id="ttot">รวม</button></div>
    </div>
    <div class="chart-box"><canvas id="barChart"></canvas></div>
    </div>

    <div class="card" id="pct-chart-card">
    <h3>📈 ร้อยละผลงานรายอำเภอ (เรียงจากมากไปน้อย)</h3>
    <div class="chart-box"><canvas id="pctChart"></canvas></div>
    </div>
  </div>

  <div id="district-charts" style="display:none;">
    <div class="da-grid">
    <div class="card" style="margin:0;background:var(--bg);border-color:var(--bdr);"><h3>📊 สัดส่วนผลงาน</h3><div style="max-width:300px;margin:0 auto;"><canvas id="daDoughnut"></canvas></div></div>
    <div class="card" style="margin:0;background:var(--bg);border-color:var(--bdr);"><h3>📊 ผลงานรายหน่วยบริการ</h3><div class="chart-box" style="height:350px;"><canvas id="daBar"></canvas></div></div>
    </div>
  </div>
</div>

<div class="card" id="table-section"><div class="card" id="table-section">
<div class="tbl-header">
<h3 id="table-section-title">📋 สรุปรายเครือข่าย (คลิก + ขยายดูรายหน่วยบริการ)</h3>
<div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
<input type="text" class="search" placeholder="🔍 ค้นหาอำเภอ / หน่วยบริการ..." oninput="filterTbl(this.value)">
<button class="btn" onclick="exportCSV('summary')">📥 ส่งออก CSV</button>
</div>
</div>
<div class="twrap">
<table id="dtbl"><thead><tr>
<th></th><th onclick="srt(1)">อำเภอ</th><th onclick="srt(2)">เครือข่าย</th><th onclick="srt(3)">ลูกข่าย</th><th onclick="srt(4)">เป้าหมาย</th><th onclick="srt(5)">ฉีดแล้ว</th><th onclick="srt(6)">% ผลงาน</th>
</tr></thead><tbody id="tbody"></tbody></table>
</div>
</div>

<footer class="footer">
<p>ข้อมูลจากระบบรายงาน สปสช. เขต 7 · จัดทำโดยระบบอัตโนมัติ</p>
<p style="margin-top:4px;">อัพเดทล่าสุด: {utime}</p>
</footer>
</div>

<button id="scrollTop" onclick="window.scrollTo(0,0)">↑</button>

<script>
try{{Chart.register(ChartDataLabels);}}catch(e){{console.warn('ChartDataLabels not loaded',e);}}
const DL={json.dumps(dl, ensure_ascii=False)};
const V11T={json.dumps(dv11t)};
const V11V={json.dumps(dv11v)};
const V12T={json.dumps(dv12t)};
const V12V={json.dumps(dv12v)};
const DPCT={json.dumps(dpct)};
const SL={json.dumps(sorted_labels, ensure_ascii=False)};
const SP={json.dumps(sorted_pcts)};
const ND={nj};

// --- Loader: hide immediately when DOM ready, fallback timeout ---
function hideLoader(){{const l=document.getElementById('loader');if(l)l.classList.add('hide');}}
if(document.readyState==='complete'||document.readyState==='interactive'){{setTimeout(hideLoader,300);}}
else{{document.addEventListener('DOMContentLoaded',()=>setTimeout(hideLoader,300));}}
setTimeout(hideLoader,3000);// safety fallback

// --- Count animation ---
// --- Active Nav ---
window.addEventListener('scroll',()=>{{{{
const btn=document.getElementById('scrollTop');
btn.classList.toggle('show',window.scrollY>400);
const sections=['district-selector','kpi-section','charts-section','table-section'];
const links=document.querySelectorAll('.nav a');
let current='';
sections.forEach(id=>{{{{const el=document.getElementById(id);if(el&&window.scrollY>=el.offsetTop-100)current=id;}}}});
links.forEach(a=>{{{{a.classList.toggle('active',a.getAttribute('href')==='#'+current);}}}});
}}}});

// --- Count animation ---
function animateValue(el, target, suffix) {{{{
  const isFloat = suffix === '%';
  let current = 0;
  const step = target / 40;
  if(target === 0) {{{{ el.textContent = "0" + suffix; return; }}}}
  const timer = setInterval(() => {{{{
    current += step;
    if(current >= target) {{{{ current = target; clearInterval(timer); }}}}
    el.textContent = (isFloat ? current.toFixed(2) : Math.round(current).toLocaleString()) + suffix;
  }}}}, 30);
}}}}

function updateKPIs(tt, tv, pct, v11t, v12t, v11v, v12v, num1, num2, lbl1, lbl2, lbl3, lbl4) {{{{
  animateValue(document.getElementById('kpi-tt'), tt, '');
  document.getElementById('kpi-sub-t').textContent = `V11: ${{v11t.toLocaleString()}} · V12: ${{v12t.toLocaleString()}}`;
  
  animateValue(document.getElementById('kpi-tv'), tv, '');
  document.getElementById('kpi-sub-v').textContent = `V11: ${{v11v.toLocaleString()}} · V12: ${{v12v.toLocaleString()}}`;
  
  animateValue(document.getElementById('kpi-pct'), parseFloat(pct), '%');
  const pfill = document.getElementById('kpi-pfill');
  pfill.style.width = Math.min(pct, 100) + '%';
  pfill.style.background = pct>=80?'linear-gradient(135deg,var(--green),var(--cyan))':pct>=50?'linear-gradient(135deg,var(--amber),#f97316)':'linear-gradient(135deg,var(--red),#f97316)';
  
  document.getElementById('kpi-lbl-units').textContent = lbl1;
  animateValue(document.getElementById('kpi-units'), num1, '');
  document.getElementById('kpi-sub-units').textContent = lbl2;
  
  document.getElementById('kpi-lbl-total-units').textContent = lbl3;
  animateValue(document.getElementById('kpi-total-units'), num2, '');
  document.getElementById('kpi-sub-total-units').textContent = lbl4;
}}}}

// --- Bar Chart ---
const barCtx=document.getElementById('barChart').getContext('2d');
let barChart=new Chart(barCtx,{{type:'bar',data:{{labels:DL,datasets:[
{{label:'เป้าหมาย V11',data:V11T,backgroundColor:'rgba(59,130,246,.7)',borderColor:'rgba(59,130,246,1)',borderWidth:1,borderRadius:4}},
{{label:'ฉีดแล้ว V11',data:V11V,backgroundColor:'rgba(16,185,129,.7)',borderColor:'rgba(16,185,129,1)',borderWidth:1,borderRadius:4}}
]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{family:"'Noto Sans Thai'"}}}}}},datalabels:{{display:false}}}},scales:{{x:{{ticks:{{color:'#94a3b8',font:{{family:"'Noto Sans Thai'",size:11}},maxRotation:45}},grid:{{color:'rgba(51,65,85,.3)'}}}},y:{{ticks:{{color:'#94a3b8',font:{{family:"'Noto Sans Thai'"}}}},grid:{{color:'rgba(51,65,85,.3)'}}}}}}}}}});

function swChart(m){{
document.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));
let t,v,lt,lv;
if(m==='v11'){{t=V11T;v=V11V;lt='เป้าหมาย V11';lv='ฉีดแล้ว V11';document.getElementById('tv11').classList.add('active');}}
else if(m==='v12'){{t=V12T;v=V12V;lt='เป้าหมาย V12';lv='ฉีดแล้ว V12';document.getElementById('tv12').classList.add('active');}}
else{{t=V11T.map((x,i)=>x+V12T[i]);v=V11V.map((x,i)=>x+V12V[i]);lt='เป้าหมายรวม';lv='ฉีดรวม';document.getElementById('ttot').classList.add('active');}}
barChart.data.datasets[0].data=t;barChart.data.datasets[0].label=lt;
barChart.data.datasets[1].data=v;barChart.data.datasets[1].label=lv;
barChart.update();
}}

// --- Pct Chart (sorted desc) ---
const pctCtx=document.getElementById('pctChart').getContext('2d');
const pctColors=SP.map(p=>p>=80?'rgba(16,185,129,.8)':p>=50?'rgba(245,158,11,.8)':'rgba(239,68,68,.8)');
const pctBorders=SP.map(p=>p>=80?'rgba(16,185,129,1)':p>=50?'rgba(245,158,11,1)':'rgba(239,68,68,1)');
let pctChart=new Chart(pctCtx,{{type:'bar',data:{{labels:SL,datasets:[{{label:'% ผลงาน',data:SP,backgroundColor:pctColors,borderColor:pctBorders,borderWidth:1,borderRadius:4}}]}},
options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},datalabels:{{anchor:'end',align:'end',color:'#f1f5f9',font:{{family:"'Noto Sans Thai'",size:11,weight:600}},formatter:v=>v.toFixed(1)+'%'}}}},
scales:{{x:{{max:Math.max(...SP)+20,ticks:{{color:'#94a3b8',callback:v=>v+'%',font:{{family:"'Noto Sans Thai'"}}}},grid:{{color:'rgba(51,65,85,.3)'}}}},y:{{ticks:{{color:'#94a3b8',font:{{family:"'Noto Sans Thai'",size:12}}}},grid:{{display:false}}}}}}}}}});

// --- Table ---
function pb(pct){{const c=pct>=80?'ph':pct>=50?'pm':'pl';return `<span class="pct-b ${{c}}">${{pct.toFixed(1)}}%</span>`;}}

function buildTbl(){{
const tb=document.getElementById('tbody');tb.innerHTML='';
ND.forEach((n,i)=>{{
const tr=document.createElement('tr');tr.classList.add('mu');
tr.dataset.district=n.district;tr.dataset.name=n.mn;
tr.innerHTML=`<td><button class="eb" onclick="tog('n${{i}}',this)">+</button></td><td>${{n.district}}</td><td>${{n.mn}}</td><td>${{n.sc}}</td><td>${{n.tt.toLocaleString()}}</td><td>${{n.tv.toLocaleString()}}</td><td>${{pb(n.pct)}}</td>`;
tb.appendChild(tr);
n.units.forEach(u=>{{
const sr=document.createElement('tr');sr.classList.add('sr','n'+i);
sr.dataset.district=n.district;sr.dataset.name=u.name;
const icon=u.type==='แม่ข่าย'?'🏥':'🏠';
sr.innerHTML=`<td></td><td>${{icon}} ${{u.name}} <span style="color:var(--fg3);font-size:.75rem">(${{u.code}})</span></td><td>${{u.type}}</td><td></td><td>${{u.tt.toLocaleString()}}</td><td>${{u.tv.toLocaleString()}}</td><td>${{pb(u.pct)}}</td>`;
tb.appendChild(sr);
}});
}});
}}

function tog(cls,btn){{
const rows=document.querySelectorAll('.'+cls);
const ex=btn.classList.toggle('ex');
btn.textContent=ex?'−':'+';
rows.forEach(r=>r.classList.toggle('show'));
}}

function filterTbl(q){{
const ql=q.toLowerCase();
document.querySelectorAll('#tbody tr').forEach(r=>{{
const d=(r.dataset.district||'').toLowerCase();
const n=(r.dataset.name||'').toLowerCase();
const m=d.includes(ql)||n.includes(ql);
if(r.classList.contains('mu'))r.style.display=m?'':'none';
else r.style.display=m&&r.classList.contains('show')?'':'none';
}});
}}

let sortDir={{}};
function srt(ci){{
const tb=document.getElementById('tbody');
const mains=[...tb.querySelectorAll('tr.mu')];
sortDir[ci]=!(sortDir[ci]||false);const dir=sortDir[ci]?1:-1;
document.querySelectorAll('thead th').forEach(th=>{{th.classList.remove('sa','sd');}});
document.querySelectorAll('thead th')[ci].classList.add(dir===1?'sa':'sd');
const groups=mains.map(m=>{{const subs=[];let s=m.nextElementSibling;while(s&&s.classList.contains('sr')){{subs.push(s);s=s.nextElementSibling;}}return{{main:m,subs}};}});
groups.sort((a,b)=>{{const av=a.main.children[ci].textContent.trim();const bv=b.main.children[ci].textContent.trim();const an=parseFloat(av.replace(/[^0-9.-]/g,''));const bn=parseFloat(bv.replace(/[^0-9.-]/g,''));if(!isNaN(an)&&!isNaN(bn))return(an-bn)*dir;return av.localeCompare(bv,'th')*dir;}});
tb.innerHTML='';groups.forEach(g=>{{tb.appendChild(g.main);g.subs.forEach(s=>tb.appendChild(s));}});
}}

// --- CSV Export ---
function exportCSV(type){{
let csv=String.fromCharCode(0xFEFF);// BOM for Excel
if(type==='summary'){{
csv+='อำเภอ,เครือข่าย,ลูกข่าย,เป้าหมาย,ฉีดแล้ว,%ผลงาน'+String.fromCharCode(10);
ND.forEach(n=>{{csv+='"'+n.district+'","'+n.mn+'",'+n.sc+','+n.tt+','+n.tv+','+n.pct+String.fromCharCode(10);}});
}}else{{
const sel=selectedDistrict;
if(!sel){{alert('กรุณาคลิกเลือกข้อมูลในตาราง Grid ก่อนดาวน์โหลดข้อมูล');return;}}
if(sel==='ภาพรวมจังหวัด'){{
csv+='รายละเอียด: ภาพรวมจังหวัดกาฬสินธุ์'+String.fromCharCode(10);
csv+='เครือข่าย (แม่ข่าย),เป้าหมายรวม,ฉีดแล้ว,%ผลงาน'+String.fromCharCode(10);
ND.forEach(n=>{{
csv+='"'+n.mn+'",'+n.tt+','+n.tv+','+n.pct+String.fromCharCode(10);
}});
}}else{{
csv+='รายละเอียดอำเภอ: '+sel+String.fromCharCode(10);
csv+='ประเภท,รหัส,ชื่อหน่วยบริการ,เป้าหมาย V11,ฉีด V11,เป้าหมาย V12,ฉีด V12,เป้าหมายรวม,ฉีดรวม,%ผลงาน'+String.fromCharCode(10);
ND.filter(n=>n.district===sel).forEach(n=>{{
n.units.forEach(u=>{{csv+='"'+u.type+'","'+u.code+'","'+u.name+'",'+u.v11t+','+u.v11v+','+u.v12t+','+u.v12v+','+u.tt+','+u.tv+','+u.pct+String.fromCharCode(10);}});
}});
}}
}}
const blob=new Blob([csv],{{type:'text/csv;charset=utf-8;'}});
const a=document.createElement('a');a.href=URL.createObjectURL(blob);
a.download=type==='summary'?'สรุปเครือข่าย_กาฬสินธุ์.csv':('รายละเอียด_'+selectedDistrict+'.csv');
a.click();
}}

// --- Theme Selector ---
function setTheme(name) {{
  document.body.className = 'theme-' + name;
  localStorage.setItem('df-theme', name);
  
  document.querySelectorAll('.theme-btn').forEach(btn => {{
    btn.classList.toggle('active', btn.dataset.theme === name);
  }});
  
  const isLight = name === 'light';
  const isBlue = name === 'blue';
  const isEmerald = name === 'emerald';
  
  let textColor = isLight ? '#334155' : isBlue ? '#8b9bb4' : isEmerald ? '#a7f3d0' : '#94a3b8';
  let gridColor = isLight ? 'rgba(226, 232, 240, 0.8)' : isBlue ? 'rgba(58, 80, 107, 0.3)' : isEmerald ? 'rgba(15, 118, 110, 0.3)' : 'rgba(51, 65, 85, 0.2)';
  
  [barChart, pctChart, daDoughnutChart, daBarChart].forEach(chart => {{
    if (chart) {{
      if (chart.options.scales) {{
        if (chart.options.scales.x) {{
          if (chart.options.scales.x.ticks) chart.options.scales.x.ticks.color = textColor;
          if (chart.options.scales.x.grid) chart.options.scales.x.grid.color = gridColor;
        }}
        if (chart.options.scales.y) {{
          if (chart.options.scales.y.ticks) chart.options.scales.y.ticks.color = textColor;
          if (chart.options.scales.y.grid) chart.options.scales.y.grid.color = gridColor;
        }}
      }}
      if (chart.options.plugins && chart.options.plugins.legend && chart.options.plugins.legend.labels) {{
        chart.options.plugins.legend.labels.color = textColor;
      }}
      chart.update('none');
    }}
  }});
}}

// --- District Analysis ---
let daDoughnutChart=null,daBarChart=null;
let selectedDistrict='';

function selectDistrictCard(name, scroll = true) {{
selectedDistrict=name;
document.querySelectorAll('.da-grid-card').forEach(c=>c.classList.remove('active'));
const activeCard=document.getElementById('dg-'+name);
if(activeCard)activeCard.classList.add('active');
showDistrict(name, scroll);
}}

function showDistrict(name, scroll = true){{
const bName=document.getElementById('daDistrictName');
if(!name) return;

bName.textContent = name === 'ภาพรวมจังหวัด' ? 'ภาพรวมจังหวัดกาฬสินธุ์' : 'อำเภอ' + name;

let tV11T=0,tV11V=0,tV12T=0,tV12V=0,tUnits=0;
let tTT=0,tTV=0,tPct=0;
const chartLabels=[];
const chartTargets=[];
const chartVaccs=[];

if(name==='ภาพรวมจังหวัด'){{
  tV11T=V11T.reduce((a,b)=>a+b,0);
  tV11V=V11V.reduce((a,b)=>a+b,0);
  tV12T=V12T.reduce((a,b)=>a+b,0);
  tV12V=V12V.reduce((a,b)=>a+b,0);
  tTT=tV11T+tV12T;
  tTV=tV11V+tV12V;
  tPct=tTT>0?((tTV/tTT)*100).toFixed(2):0;
  
  document.getElementById('provincial-charts').style.display = 'block';
  document.getElementById('district-charts').style.display = 'none';
  
  updateKPIs(tTT, tTV, tPct, tV11T, tV12T, tV11V, tV12V, {len(districts)}, {tu}, '🏥 อำเภอ', 'อำเภอ', '🏠 หน่วยบริการ', 'แม่ข่าย+ลูกข่าย');

  const tblHead = document.querySelector('#dtbl thead');
  const titleEl = document.getElementById('table-section-title');
  if (titleEl) titleEl.innerHTML = '📋 สรุปรายเครือข่าย (คลิก + ขยายดูรายหน่วยบริการ)';
  if (tblHead) tblHead.innerHTML = '<tr><th></th><th onclick="srt(1)">อำเภอ</th><th onclick="srt(2)">เครือข่าย</th><th onclick="srt(3)">ลูกข่าย</th><th onclick="srt(4)">เป้าหมาย</th><th onclick="srt(5)">ฉีดแล้ว</th><th onclick="srt(6)">% ผลงาน</th></tr>';
  buildTbl();
}} else {{
  const nets = ND.filter(n => n.district === name);
  let dUnits = 0;
  nets.forEach(n => {{
    tV11T += n.v11t;
    tV11V += n.v11v;
    tV12T += n.v12t;
    tV12V += n.v12v;
    n.units.forEach(u => {{
      dUnits++;
      chartLabels.push(u.name);
      chartTargets.push(u.tt);
      chartVaccs.push(u.tv);
    }});
  }});
  tTT=tV11T+tV12T;
  tTV=tV11V+tV12V;
  tPct=tTT>0?((tTV/tTT)*100).toFixed(2):0;

  document.getElementById('provincial-charts').style.display = 'none';
  document.getElementById('district-charts').style.display = 'block';

  updateKPIs(tTT, tTV, tPct, tV11T, tV12T, tV11V, tV12V, nets.length, dUnits, '🏥 เครือข่าย', 'แม่ข่าย', '🏠 หน่วยบริการ', 'ลูกข่ายในอำเภอ');

  if(daDoughnutChart)daDoughnutChart.destroy();
  const dCtx=document.getElementById('daDoughnut').getContext('2d');
  daDoughnutChart=new Chart(dCtx,{{type:'doughnut',data:{{labels:['ฉีดแล้ว','ยังไม่ฉีด'],datasets:[{{data:[tTV,Math.max(tTT-tTV,0)],backgroundColor:['rgba(16,185,129,.8)','rgba(51,65,85,.5)'],borderWidth:0}}]}},
  options:{{responsive:true,cutout:'65%',plugins:{{legend:{{labels:{{color: document.body.classList.contains('theme-light') ? '#334155' : '#94a3b8', font:{{family:"\'Noto Sans Thai\'"}}}}}},datalabels:{{color:'#f1f5f9',font:{{family:"\'Noto Sans Thai\'",size:13,weight:600}},formatter:(v,ctx)=>{{const t=ctx.dataset.data.reduce((a,b)=>a+b,0);return t>0?(v/t*100).toFixed(1)+'%':'0%';}}}}}}}}}});

  if(daBarChart)daBarChart.destroy();
  const bCtx=document.getElementById('daBar').getContext('2d');
  
  const textColor = document.body.classList.contains('theme-light') ? '#334155' : '#94a3b8';
  const gridColor = document.body.classList.contains('theme-light') ? 'rgba(226, 232, 240, 0.8)' : 'rgba(51, 65, 85, 0.2)';

  daBarChart=new Chart(bCtx,{{type:'bar',data:{{labels:chartLabels,datasets:[
  {{label:'เป้าหมาย',data:chartTargets,backgroundColor:'rgba(59,130,246,.6)',borderRadius:3}},
  {{label:'ฉีดแล้ว',data:chartVaccs,backgroundColor:'rgba(16,185,129,.6)',borderRadius:3}}
  ]}},options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{labels:{{color: textColor, font:{{family:"\'Noto Sans Thai\'"}}}}}},datalabels:{{display:false}}}},scales:{{x:{{ticks:{{color: textColor, font:{{family:"\'Noto Sans Thai\'"}}}},grid:{{color: gridColor}}}},y:{{ticks:{{color: textColor, font:{{family:"\'Noto Sans Thai\'",size:10}}}},grid:{{display:false}}}}}}}}}});

  const tb = document.getElementById('tbody');
  const titleEl = document.getElementById('table-section-title');
  const tblHead = document.querySelector('#dtbl thead');
  if (titleEl) titleEl.innerHTML = '📋 รายละเอียดหน่วยบริการในพื้นที่ อำเภอ' + name;
  if (tblHead) tblHead.innerHTML = '<tr><th>ลำดับ</th><th>ประเภท</th><th>รหัส</th><th style="text-align:left;">ชื่อหน่วยบริการ</th><th>เป้าหมาย</th><th>ฉีดแล้ว</th><th>% ผลงาน</th></tr>';
  tb.innerHTML = '';
  let idx = 1;
  nets.forEach(n => {{
    n.units.forEach(u => {{
      const tr = document.createElement('tr');
      if (u.type === 'แม่ข่าย') tr.classList.add('mu');
      const icon = u.type === 'แม่ข่าย' ? '🏥 แม่ข่าย' : '🏠 ลูกข่าย';
      tr.dataset.district = name;
      tr.dataset.name = u.name;
      tr.innerHTML = '<td>' + idx++ + '</td><td>' + icon + '</td><td>' + u.code + '</td><td style="text-align:left;">' + u.name + '</td><td>' + u.tt.toLocaleString() + '</td><td>' + u.tv.toLocaleString() + '</td><td>' + pb(u.pct) + '</td>';
      tb.appendChild(tr);
    }});
  }});
}}

const currentTheme = localStorage.getItem('df-theme') || 'dark';
setTheme(currentTheme);

if (scroll) {{
  document.getElementById('district-selector').scrollIntoView({{behavior:'smooth',block:'start'}});
}}
}}

buildTbl();
const dfTheme = localStorage.getItem('df-theme') || 'dark';
setTheme(dfTheme);
selectDistrictCard('ภาพรวมจังหวัด', false); // provincial overview on load without scroll
</script>
</body>
</html>'''

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"📈 สร้าง Dashboard เรียบร้อย: {output_path}")

def find_git_exe():
    git_path = shutil.which('git')
    if git_path:
        return git_path
    standard_paths = [
        r"C:\Program Files\Git\cmd\git.exe",
        r"C:\Program Files (x86)\Git\cmd\git.exe",
    ]
    userprofile = os.environ.get('USERPROFILE', '')
    if userprofile:
        github_desktop_dir = os.path.join(userprofile, r"AppData\Local\GitHubDesktop")
        if os.path.exists(github_desktop_dir):
            for root, dirs, files in os.walk(github_desktop_dir):
                if 'git.exe' in files:
                    return os.path.join(root, 'git.exe')
    for p in standard_paths:
        if os.path.exists(p):
            return p
    return "git"

def find_repo_dir(repo_name):
    # Check script directory and parent directory first (for GitHub Actions / portable runs)
    if os.path.isdir(os.path.join(SCRIPT_DIR, '.git')):
        return SCRIPT_DIR
    parent_dir = os.path.dirname(SCRIPT_DIR)
    if os.path.isdir(os.path.join(parent_dir, '.git')):
        return parent_dir
        
    paths_to_check = [
        os.path.join(r"C:\GitRepos", repo_name),
    ]
    userprofile = os.environ.get('USERPROFILE', '')
    if userprofile:
        paths_to_check.extend([
            os.path.join(userprofile, "OneDrive", "เอกสาร", "GitHub", repo_name),
            os.path.join(userprofile, "OneDrive", "Documents", "GitHub", repo_name),
            os.path.join(userprofile, "Documents", "GitHub", repo_name),
            os.path.join(userprofile, repo_name),
            os.path.join(r"C:\Users\zole3", repo_name),
        ])
    for p in paths_to_check:
        if os.path.exists(p) and os.path.isdir(os.path.join(p, '.git')):
            return p
    return None

def deploy_to_github(html_output):
    print("\n🚀 กำลังส่งออกข้อมูลขึ้น GitHub ที่โฟลเดอร์ Repository: influenza")
    git_exe = find_git_exe()
    git_dir = find_repo_dir('influenza')
    if not git_dir:
        print("❌ ไม่พบโฟลเดอร์ Git Repository ของ influenza ในพาธที่กำหนด")
        print("💡 คุณสามารถทำการคัดลอกไฟล์ Fludashboard.html ไปยัง Git Repo ด้วยตนเอง")
        return

    # Copy files
    try:
        dest_flu = os.path.join(git_dir, 'Fludashboard.html')
        dest_idx = os.path.join(git_dir, 'index.html')
        if os.path.abspath(html_output) != os.path.abspath(dest_flu):
            shutil.copy2(html_output, dest_flu)
        if os.path.abspath(html_output) != os.path.abspath(dest_idx):
            shutil.copy2(html_output, dest_idx)
        print(f"✅ คัดลอกแดชบอร์ดไปยัง Repository เรียบร้อย! ({git_dir})")
    except Exception as e:
        print(f"❌ ไม่สามารถคัดลอกไฟล์ไปยัง Git Repo: {e}")
        return

    # Run git commands
    try:
        print("Checking git status...")
        subprocess.run([git_exe, 'status'], cwd=git_dir)
        print("Executing: git add Fludashboard.html index.html")
        subprocess.run([git_exe, 'add', 'Fludashboard.html', 'index.html'], cwd=git_dir, check=True)
        utime = datetime.now().strftime('%d/%m/%Y %H:%M น.')
        commit_msg = f"Auto-update Influenza Dashboard (ข้อมูล ณ วันที่ {utime})"
        print(f"Executing: git commit -m \"{commit_msg}\"")
        subprocess.run([git_exe, 'commit', '-m', commit_msg], cwd=git_dir, capture_output=True)
        print("Executing: git push origin main")
        res = subprocess.run([git_exe, 'push', 'origin', 'main'], cwd=git_dir, capture_output=True, text=True, encoding='utf-8', errors='ignore')
        print(res.stdout)
        print(res.stderr)
        print("🎉 อัปเดทข้อมูลขึ้น GitHub เรียบร้อยแล้ว!")
        print("🔗 ลิงก์แดชบอร์ด: https://zole3500-ana.github.io/influenza/index.html")
    except Exception as e:
        print(f"❌ เกิดข้อผิดพลาดในการดำเนินการ Git: {e}")

# ============================================================
# MAIN
# ============================================================

def main():
    print("="*60)
    print("  ระบบอัพเดทรายงานฉีดวัคซีนไข้หวัดใหญ่ จ.กาฬสินธุ์")
    print("="*60)
    print()
    if not os.path.exists(INPUT_FILE):
        print(f"❌ ไม่พบไฟล์: {INPUT_FILE}")
        sys.exit(1)
    rows = load_data(INPUT_FILE)
    units = pivot_data(rows)
    networks = aggregate_by_network(units)
    districts = aggregate_by_district(networks)
    print(f"\n📋 สรุป: {len(districts)} อำเภอ, {len(networks)} เครือข่าย, {len(units)} หน่วยบริการ\n")
    generate_excel(units, networks, districts, EXCEL_OUTPUT)
    generate_dashboard(units, networks, districts, HTML_OUTPUT)
    
    # Auto deploy to GitHub
    deploy_to_github(HTML_OUTPUT)
    
    print(f"\n✅ เสร็จสิ้น!")
    print(f"   📊 {EXCEL_OUTPUT}")
    print(f"   📈 {HTML_OUTPUT}")
    print(f"\n🔄 อัพเดทครั้งต่อไป: uv run --with openpyxl --no-project update_report.py")

if __name__ == '__main__':
    main()
